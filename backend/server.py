from collections import defaultdict
from datetime import datetime, timedelta
import csv
import io
import logging
import os
from math import atan2, cos, radians, sin, sqrt
from pathlib import Path
from typing import Any, Dict, List, Literal, Optional

from dotenv import load_dotenv
from fastapi import APIRouter, Depends, FastAPI, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, EmailStr, validator
from starlette.middleware.cors import CORSMiddleware

from auth_helper import get_current_user
from database import db

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

app = FastAPI(title="GeoPunch API", version="4.0.0")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


class WorkdaysConfig(BaseModel):
    monday: bool = True
    tuesday: bool = True
    wednesday: bool = True
    thursday: bool = True
    friday: bool = True
    saturday: bool = False
    sunday: bool = False


class ScheduleConfig(BaseModel):
    startTime: str = "09:00"
    endTime: str = "18:00"
    marginMinutes: int = 120


class WorkplaceCreate(BaseModel):
    name: str
    latitude: float
    longitude: float
    radiusMeters: int = 150
    workdays: WorkdaysConfig = WorkdaysConfig()
    schedule: Optional[ScheduleConfig] = None

    @validator("radiusMeters")
    def validate_radius(cls, value: int) -> int:
        if value < 50 or value > 300:
            raise ValueError("Raio deve estar entre 50m e 300m")
        return value

    @validator("name")
    def validate_name(cls, value: str) -> str:
        if not value or not value.strip():
            raise ValueError("Nome é obrigatório")
        return value.strip()


class WorkplaceUpdate(BaseModel):
    name: Optional[str] = None
    radiusMeters: Optional[int] = None
    workdays: Optional[WorkdaysConfig] = None
    schedule: Optional[ScheduleConfig] = None

    @validator("radiusMeters")
    def validate_radius(cls, value: Optional[int]) -> Optional[int]:
        if value is not None and (value < 50 or value > 300):
            raise ValueError("Raio deve estar entre 50m e 300m")
        return value


class WorkplaceResponse(BaseModel):
    id: str
    name: str
    latitude: float
    longitude: float
    radiusMeters: int
    workdays: Dict[str, bool]
    schedule: Optional[Dict[str, Any]] = None
    locationLocked: bool = True
    configuredAt: datetime
    isActive: bool = False
    createdAt: datetime
    contextType: Literal["personal", "enterprise"]
    enterpriseId: Optional[str] = None
    assignmentId: Optional[str] = None
    assignedEmployeeCount: Optional[int] = None


class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    employeeId: Optional[str] = None
    role: str
    accountType: str
    enterpriseId: Optional[str] = None
    enterpriseName: Optional[str] = None
    activeWorkplaceId: Optional[str] = None
    createdAt: datetime


class EnterpriseBootstrapRequest(BaseModel):
    name: str
    nif: Optional[str] = None

    @validator("name")
    def validate_name(cls, value: str) -> str:
        if not value or not value.strip():
            raise ValueError("Nome da empresa é obrigatório")
        return value.strip()


class EnterpriseResponse(BaseModel):
    id: str
    name: str
    nif: Optional[str] = None
    ownerUserId: str
    createdAt: datetime


class EnterpriseInvitationCreate(BaseModel):
    email: EmailStr


class EnterpriseMembershipResponse(BaseModel):
    id: str
    enterpriseId: str
    email: str
    userId: Optional[str] = None
    userName: Optional[str] = None
    status: str
    invitedBy: str
    acceptedAt: Optional[datetime] = None
    respondedAt: Optional[datetime] = None
    createdAt: datetime
    assignedWorkplaceIds: List[str] = []


class AssignEnterpriseWorkplaceRequest(BaseModel):
    employeeUserId: str
    workplaceId: str


class EmployeeWorkplaceAssignmentResponse(BaseModel):
    id: str
    enterpriseId: str
    employeeUserId: str
    workplaceId: str
    assignedBy: str
    createdAt: datetime


class PunchCreate(BaseModel):
    punchType: Literal["IN", "OUT", "BREAK_START", "BREAK_END"]
    latitude: float
    longitude: float
    accuracy: float
    deviceTime: Optional[datetime] = None
    note: Optional[str] = None
    method: Literal["manual", "geofence_suggestion"] = "manual"


class PunchResponse(BaseModel):
    id: str
    userId: str
    workplaceId: str
    workplaceName: str
    date: str
    punchType: str
    occurredAt: datetime
    receivedAt: datetime
    latitude: float
    longitude: float
    accuracyMeters: float
    distanceToWorkplaceMeters: float
    method: str
    outsideWorkplace: bool
    note: Optional[str] = None


class GeofenceEventCreate(BaseModel):
    eventId: str
    eventType: Literal["ENTER", "EXIT"]
    latitude: float
    longitude: float
    accuracy: float
    deviceTime: Optional[datetime] = None


class DayTimesheetResponse(BaseModel):
    date: str
    workplaceName: str
    workplaceId: str
    isScheduledWorkday: bool
    punches: List[Dict[str, Any]]
    grossMinutes: int = 0
    breakMinutes: int = 0
    netWorkedMinutes: int = 0
    netWorkedFormatted: str = "00:00"
    status: str
    anomalies: List[str] = []


def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    earth_radius_m = 6371000
    lat1_rad = radians(lat1)
    lat2_rad = radians(lat2)
    delta_lat = radians(lat2 - lat1)
    delta_lon = radians(lon2 - lon1)
    a = sin(delta_lat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(delta_lon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return earth_radius_m * c


def format_minutes(minutes: int) -> str:
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def get_today_date() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


def is_workday(workdays: Dict[str, bool], date: datetime) -> bool:
    day_map = {
        0: "monday",
        1: "tuesday",
        2: "wednesday",
        3: "thursday",
        4: "friday",
        5: "saturday",
        6: "sunday",
    }
    return workdays.get(day_map[date.weekday()], False)


def generate_maps_link(lat: float, lng: float) -> str:
    return f"https://maps.google.com/?q={lat},{lng}"


def parse_db_datetime(value: Optional[str]) -> Optional[datetime]:
    return datetime.fromisoformat(value) if value else None


def require_enterprise_owner(user: Dict[str, Any]) -> None:
    if user.get("role") != "enterprise_owner":
        raise HTTPException(status_code=403, detail="Acesso negado. Conta empresa necessária.")


def require_worker_account(user: Dict[str, Any]) -> None:
    if user.get("role") == "enterprise_owner":
        raise HTTPException(status_code=403, detail="Conta empresa não pode registar ponto.")


def normalize_workplace_response(
    workplace: Dict[str, Any],
    *,
    active_workplace_id: Optional[str],
    assignment: Optional[Dict[str, Any]] = None,
    assigned_employee_count: Optional[int] = None,
) -> WorkplaceResponse:
    return WorkplaceResponse(
        id=workplace["id"],
        name=workplace["name"],
        latitude=float(workplace["latitude"]),
        longitude=float(workplace["longitude"]),
        radiusMeters=workplace["radius_meters"],
        workdays=workplace.get("workdays", {}),
        schedule=workplace.get("schedule"),
        locationLocked=workplace.get("location_locked", True),
        configuredAt=parse_db_datetime(workplace.get("configured_at") or workplace.get("created_at")) or datetime.utcnow(),
        isActive=workplace["id"] == active_workplace_id if active_workplace_id else False,
        createdAt=parse_db_datetime(workplace["created_at"]) or datetime.utcnow(),
        contextType="enterprise" if workplace.get("enterprise_id") else "personal",
        enterpriseId=workplace.get("enterprise_id"),
        assignmentId=assignment.get("id") if assignment else None,
        assignedEmployeeCount=assigned_employee_count,
    )


async def get_user_enterprise(user: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    enterprise_id = user.get("enterprise_id")
    if not enterprise_id:
        return None
    return await db.find_enterprise_by_id(enterprise_id)


async def get_accessible_workplaces(user: Dict[str, Any]) -> List[Dict[str, Any]]:
    if user.get("role") == "enterprise_owner":
        if not user.get("enterprise_id"):
            return []
        return await db.find_enterprise_workplaces(user["enterprise_id"])

    if user.get("enterprise_id"):
        return await db.find_employee_assigned_workplaces(user["enterprise_id"], user["id"])

    return await db.find_personal_workplaces_by_user(user["id"])


async def ensure_workplace_access(user: Dict[str, Any], workplace_id: str, *, for_management: bool = False) -> Dict[str, Any]:
    workplace = await db.find_workplace_by_id(workplace_id)
    if not workplace:
        raise HTTPException(status_code=404, detail="Local de trabalho não encontrado")

    if user.get("role") == "enterprise_owner":
        if workplace.get("enterprise_id") != user.get("enterprise_id"):
            raise HTTPException(status_code=403, detail="Sem acesso a este local de trabalho")
        return workplace

    if user.get("enterprise_id"):
        if for_management:
            raise HTTPException(status_code=403, detail="Funcionário não pode gerir locais da empresa")
        assignment = await db.find_employee_workplace_assignment(user["enterprise_id"], user["id"], workplace_id)
        if not assignment:
            raise HTTPException(status_code=403, detail="Sem acesso a este local de trabalho")
        workplace["assignment"] = assignment
        return workplace

    if workplace.get("user_id") != user["id"] or workplace.get("enterprise_id"):
        raise HTTPException(status_code=403, detail="Sem acesso a este local de trabalho")

    return workplace


def group_punches_by_day(punches_data: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    days: Dict[str, Dict[str, Any]] = {}
    for punch in punches_data:
        date_key = punch["date"]
        if date_key not in days:
            days[date_key] = {
                "workplaceId": punch["workplace_id"],
                "workplaceName": punch["workplace_name"],
                "punches": [],
                "anomalies": [],
            }

        occurred_at = parse_db_datetime(punch["occurred_at"]) or datetime.utcnow()
        days[date_key]["punches"].append(
            {
                "id": punch["id"],
                "type": punch["punch_type"],
                "occurredAt": occurred_at,
                "method": punch["method"],
                "outsideWorkplace": punch["outside_workplace"],
                "distance": float(punch["distance_to_workplace_meters"]),
                "accuracy": float(punch["accuracy_meters"]),
                "note": punch.get("note"),
                "mapsLink": generate_maps_link(float(punch["latitude"]), float(punch["longitude"])),
            }
        )

        if punch["outside_workplace"]:
            days[date_key]["anomalies"].append(f"{punch['punch_type']}: fora do local")
        if float(punch["accuracy_meters"]) > 50:
            days[date_key]["anomalies"].append(
                f"{punch['punch_type']}: GPS impreciso ({int(float(punch['accuracy_meters']))}m)"
            )

    return days


def build_timesheet_rows(
    grouped_days: Dict[str, Dict[str, Any]],
    workplaces_by_id: Dict[str, Dict[str, Any]],
) -> List[DayTimesheetResponse]:
    result: List[DayTimesheetResponse] = []

    for date_str, day_data in sorted(grouped_days.items(), reverse=True):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        workplace = workplaces_by_id.get(day_data["workplaceId"])
        is_workday_flag = bool(workplace and workplace.get("workdays") and is_workday(workplace["workdays"], date_obj))

        gross_minutes = 0
        break_minutes = 0
        status = "not_started"
        punch_in = None
        punch_out = None
        current_break_start = None
        break_pairs = []

        for punch in day_data["punches"]:
            if punch["type"] == "IN":
                punch_in = punch["occurredAt"]
                status = "working"
            elif punch["type"] == "OUT":
                punch_out = punch["occurredAt"]
                status = "finished"
            elif punch["type"] == "BREAK_START":
                current_break_start = punch["occurredAt"]
                status = "on_break"
            elif punch["type"] == "BREAK_END" and current_break_start:
                break_pairs.append((current_break_start, punch["occurredAt"]))
                current_break_start = None
                status = "working"

        if punch_in:
            finish_time = punch_out or datetime.utcnow()
            gross_minutes = int((finish_time - punch_in).total_seconds() / 60)

        for start, end in break_pairs:
            break_minutes += int((end - start).total_seconds() / 60)

        if current_break_start:
            break_minutes += int((datetime.utcnow() - current_break_start).total_seconds() / 60)

        net_minutes = max(0, gross_minutes - break_minutes)

        if punch_in and not is_workday_flag:
            day_data["anomalies"].append("Dia não agendado como dia de trabalho")

        result.append(
            DayTimesheetResponse(
                date=date_str,
                workplaceName=day_data["workplaceName"],
                workplaceId=day_data["workplaceId"],
                isScheduledWorkday=is_workday_flag,
                punches=day_data["punches"],
                grossMinutes=gross_minutes,
                breakMinutes=break_minutes,
                netWorkedMinutes=net_minutes,
                netWorkedFormatted=format_minutes(net_minutes),
                status=status,
                anomalies=list(dict.fromkeys(day_data["anomalies"])),
            )
        )

    return result


async def build_user_timesheet(
    user_id: str,
    workplaces: List[Dict[str, Any]],
    from_date: str,
    to_date: str,
) -> List[DayTimesheetResponse]:
    workplace_ids = [workplace["id"] for workplace in workplaces]
    punches_data = await db.find_punches_by_user_date_range(user_id, from_date, to_date, workplace_ids or None)
    grouped_days = group_punches_by_day(punches_data)
    workplaces_by_id = {workplace["id"]: workplace for workplace in workplaces}
    return build_timesheet_rows(grouped_days, workplaces_by_id)


async def build_today_status(user: Dict[str, Any]) -> Dict[str, Any]:
    today = get_today_date()
    today_date = datetime.utcnow()
    workplace = None

    active_workplace_id = user.get("active_workplace_id")
    if active_workplace_id:
        try:
            workplace = await ensure_workplace_access(user, active_workplace_id)
        except HTTPException:
            workplace = None

    punches_data = await db.find_punches({"user_id": user["id"], "date": today})
    punches = []
    for punch in punches_data:
        punch_copy = punch.copy()
        punch_copy["occurredAt"] = parse_db_datetime(punch["occurred_at"]) or datetime.utcnow()
        punches.append(punch_copy)
    punches.sort(key=lambda item: item["occurredAt"])

    punch_data: Dict[str, Any] = {"in": None, "out": None, "breaks": []}
    current_break_start = None

    for punch in punches:
        if punch["punch_type"] == "IN":
            punch_data["in"] = punch
        elif punch["punch_type"] == "OUT":
            punch_data["out"] = punch
        elif punch["punch_type"] == "BREAK_START":
            current_break_start = punch
        elif punch["punch_type"] == "BREAK_END" and current_break_start:
            punch_data["breaks"].append({"start": current_break_start, "end": punch})
            current_break_start = None

    if current_break_start:
        punch_data["breaks"].append({"start": current_break_start, "end": None})

    gross_minutes = 0
    break_minutes = 0
    status = "not_started"

    if punch_data["in"]:
        finish_time = punch_data["out"]["occurredAt"] if punch_data["out"] else datetime.utcnow()
        gross_minutes = int((finish_time - punch_data["in"]["occurredAt"]).total_seconds() / 60)
        status = "finished" if punch_data["out"] else "working"
        if current_break_start:
            status = "on_break"

    for break_item in punch_data["breaks"]:
        end_time = break_item["end"]["occurredAt"] if break_item["end"] else datetime.utcnow()
        break_minutes += int((end_time - break_item["start"]["occurredAt"]).total_seconds() / 60)

    net_minutes = max(0, gross_minutes - break_minutes)
    is_workday_flag = bool(workplace and workplace.get("workdays") and is_workday(workplace["workdays"], today_date))

    return {
        "date": today,
        "isScheduledWorkday": is_workday_flag,
        "workplace": {
            "id": workplace["id"],
            "name": workplace["name"],
            "latitude": float(workplace["latitude"]),
            "longitude": float(workplace["longitude"]),
            "radiusMeters": workplace["radius_meters"],
            "workdays": workplace.get("workdays"),
            "schedule": workplace.get("schedule"),
            "mapsLink": generate_maps_link(float(workplace["latitude"]), float(workplace["longitude"])),
        } if workplace else None,
        "punchIn": {
            "occurredAt": punch_data["in"]["occurredAt"],
            "method": punch_data["in"]["method"],
            "outsideWorkplace": punch_data["in"]["outside_workplace"],
        } if punch_data["in"] else None,
        "punchOut": {
            "occurredAt": punch_data["out"]["occurredAt"],
            "method": punch_data["out"]["method"],
            "outsideWorkplace": punch_data["out"]["outside_workplace"],
        } if punch_data["out"] else None,
        "breaks": [
            {
                "startedAt": item["start"]["occurredAt"],
                "endedAt": item["end"]["occurredAt"] if item["end"] else None,
                "durationMinutes": int(
                    (((item["end"]["occurredAt"] if item["end"] else datetime.utcnow()) - item["start"]["occurredAt"]).total_seconds()) / 60
                ),
            }
            for item in punch_data["breaks"]
        ],
        "grossMinutes": gross_minutes,
        "breakMinutes": break_minutes,
        "netWorkedMinutes": net_minutes,
        "netWorkedFormatted": format_minutes(net_minutes),
        "status": status,
    }


async def build_membership_response(
    membership: Dict[str, Any],
    assignments_by_user_id: Dict[str, List[str]],
) -> EnterpriseMembershipResponse:
    user_name = None
    if membership.get("user_id"):
        profile = await db.find_profile_by_id(membership["user_id"])
        user_name = profile.get("name") if profile else None

    return EnterpriseMembershipResponse(
        id=membership["id"],
        enterpriseId=membership["enterprise_id"],
        email=membership["email"],
        userId=membership.get("user_id"),
        userName=user_name,
        status=membership["status"],
        invitedBy=membership["invited_by"],
        acceptedAt=parse_db_datetime(membership.get("accepted_at")),
        respondedAt=parse_db_datetime(membership.get("responded_at")),
        createdAt=parse_db_datetime(membership["created_at"]) or datetime.utcnow(),
        assignedWorkplaceIds=assignments_by_user_id.get(membership.get("user_id") or "", []),
    )


@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user=Depends(get_current_user)):
    enterprise = await get_user_enterprise(user)
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user["name"],
        employeeId=user.get("employee_id"),
        role=user.get("role", "personal_user"),
        accountType=user.get("account_type", "personal"),
        enterpriseId=user.get("enterprise_id"),
        enterpriseName=enterprise.get("name") if enterprise else None,
        activeWorkplaceId=user.get("active_workplace_id"),
        createdAt=parse_db_datetime(user["created_at"]) or datetime.utcnow(),
    )


@api_router.post("/enterprise/bootstrap", response_model=EnterpriseResponse)
async def bootstrap_enterprise(payload: EnterpriseBootstrapRequest, user=Depends(get_current_user)):
    require_enterprise_owner(user)

    existing = await db.find_enterprise_by_owner(user["id"])
    if existing:
        if not user.get("enterprise_id"):
            await db.update_profile(user["id"], {"enterprise_id": existing["id"]})
        return EnterpriseResponse(
            id=existing["id"],
            name=existing["name"],
            nif=existing.get("nif"),
            ownerUserId=existing["owner_user_id"],
            createdAt=parse_db_datetime(existing["created_at"]) or datetime.utcnow(),
        )

    created = await db.create_enterprise(
        {
            "owner_user_id": user["id"],
            "name": payload.name,
            "nif": payload.nif,
            "created_at": datetime.utcnow().isoformat(),
        }
    )
    if not created:
        raise HTTPException(status_code=500, detail="Erro ao criar empresa")

    await db.update_profile(user["id"], {"enterprise_id": created["id"], "role": "enterprise_owner"})

    return EnterpriseResponse(
        id=created["id"],
        name=created["name"],
        nif=created.get("nif"),
        ownerUserId=created["owner_user_id"],
        createdAt=parse_db_datetime(created["created_at"]) or datetime.utcnow(),
    )


@api_router.get("/enterprise", response_model=Optional[EnterpriseResponse])
async def get_current_enterprise(user=Depends(get_current_user)):
    enterprise = await get_user_enterprise(user)
    if not enterprise:
        return None
    return EnterpriseResponse(
        id=enterprise["id"],
        name=enterprise["name"],
        nif=enterprise.get("nif"),
        ownerUserId=enterprise["owner_user_id"],
        createdAt=parse_db_datetime(enterprise["created_at"]) or datetime.utcnow(),
    )


@api_router.post("/enterprise/invitations", response_model=EnterpriseMembershipResponse)
async def create_enterprise_invitation(payload: EnterpriseInvitationCreate, user=Depends(get_current_user)):
    require_enterprise_owner(user)
    enterprise_id = user.get("enterprise_id")
    if not enterprise_id:
        raise HTTPException(status_code=400, detail="Conta empresa ainda não foi configurada")

    target_email = payload.email.lower()
    if target_email == user["email"].lower():
        raise HTTPException(status_code=400, detail="Não pode convidar o próprio email")

    existing_membership = await db.find_enterprise_membership_by_email(enterprise_id, target_email, ["pending", "accepted"])
    if existing_membership:
        raise HTTPException(status_code=400, detail="Já existe um convite ou associação ativa para este email")

    target_profile = await db.find_profile_by_email(target_email)
    if target_profile and target_profile.get("enterprise_id") and target_profile["enterprise_id"] != enterprise_id:
        raise HTTPException(status_code=400, detail="Este utilizador já está associado a outra empresa")

    created = await db.create_enterprise_membership(
        {
            "enterprise_id": enterprise_id,
            "user_id": target_profile["id"] if target_profile else None,
            "email": target_email,
            "invited_by": user["id"],
            "status": "pending",
            "created_at": datetime.utcnow().isoformat(),
        }
    )
    if not created:
        raise HTTPException(status_code=500, detail="Erro ao criar convite")

    return await build_membership_response(created, {})


@api_router.get("/enterprise/memberships", response_model=List[EnterpriseMembershipResponse])
async def list_enterprise_memberships(user=Depends(get_current_user)):
    require_enterprise_owner(user)
    enterprise_id = user.get("enterprise_id")
    if not enterprise_id:
        return []

    memberships = await db.list_enterprise_memberships(enterprise_id)
    assignments = await db.list_employee_workplace_assignments(enterprise_id)
    assignments_by_user_id: Dict[str, List[str]] = defaultdict(list)
    for assignment in assignments:
        assignments_by_user_id[assignment["employee_user_id"]].append(assignment["workplace_id"])

    result = []
    for membership in memberships:
        result.append(await build_membership_response(membership, assignments_by_user_id))
    return result


@api_router.get("/enterprise/invitations/mine", response_model=List[EnterpriseMembershipResponse])
async def list_my_pending_invitations(user=Depends(get_current_user)):
    if user.get("role") == "enterprise_owner":
        return []

    memberships = await db.list_pending_memberships_for_user(user["id"], user["email"])
    result = []
    for membership in memberships:
        result.append(await build_membership_response(membership, {}))
    return result


@api_router.post("/enterprise/memberships/{membership_id}/accept", response_model=EnterpriseMembershipResponse)
async def accept_membership(membership_id: str, user=Depends(get_current_user)):
    membership = await db.find_enterprise_membership_by_id(membership_id)
    if not membership or membership["status"] != "pending":
        raise HTTPException(status_code=404, detail="Convite não encontrado")

    if membership["email"].lower() != user["email"].lower():
        raise HTTPException(status_code=403, detail="Este convite não pertence ao utilizador autenticado")

    if user.get("enterprise_id") and user.get("enterprise_id") != membership["enterprise_id"]:
        raise HTTPException(status_code=400, detail="Já está associado a outra empresa")

    now_iso = datetime.utcnow().isoformat()
    updated = await db.update_enterprise_membership(
        membership_id,
        {
            "user_id": user["id"],
            "status": "accepted",
            "accepted_at": now_iso,
            "responded_at": now_iso,
        },
    )
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao aceitar convite")

    await db.update_profile(
        user["id"],
        {
            "role": "employee",
            "enterprise_id": membership["enterprise_id"],
            "active_workplace_id": None,
        },
    )

    return await build_membership_response(updated, {})


@api_router.post("/enterprise/memberships/{membership_id}/reject", response_model=EnterpriseMembershipResponse)
async def reject_membership(membership_id: str, user=Depends(get_current_user)):
    membership = await db.find_enterprise_membership_by_id(membership_id)
    if not membership or membership["status"] != "pending":
        raise HTTPException(status_code=404, detail="Convite não encontrado")

    if membership["email"].lower() != user["email"].lower():
        raise HTTPException(status_code=403, detail="Este convite não pertence ao utilizador autenticado")

    updated = await db.update_enterprise_membership(
        membership_id,
        {
            "user_id": user["id"],
            "status": "rejected",
            "responded_at": datetime.utcnow().isoformat(),
        },
    )
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao rejeitar convite")

    return await build_membership_response(updated, {})


@api_router.delete("/enterprise/memberships/{membership_id}")
async def remove_membership(membership_id: str, user=Depends(get_current_user)):
    require_enterprise_owner(user)
    membership = await db.find_enterprise_membership_by_id(membership_id)
    if not membership or membership["enterprise_id"] != user.get("enterprise_id"):
        raise HTTPException(status_code=404, detail="Associação não encontrada")

    now_iso = datetime.utcnow().isoformat()
    updated = await db.update_enterprise_membership(
        membership_id,
        {
            "status": "removed",
            "removed_at": now_iso,
            "responded_at": now_iso,
        },
    )
    if not updated:
        raise HTTPException(status_code=500, detail="Erro ao remover associação")

    if membership.get("user_id"):
        await db.update_profile(
            membership["user_id"],
            {
                "role": "personal_user",
                "enterprise_id": None,
                "active_workplace_id": None,
            },
        )
        await db.clear_employee_workplace_assignments(user["enterprise_id"], membership["user_id"])

    return {"message": "Associação removida com sucesso"}


@api_router.post("/enterprise/workplace-assignments", response_model=EmployeeWorkplaceAssignmentResponse)
async def assign_enterprise_workplace(payload: AssignEnterpriseWorkplaceRequest, user=Depends(get_current_user)):
    require_enterprise_owner(user)
    enterprise_id = user.get("enterprise_id")
    if not enterprise_id:
        raise HTTPException(status_code=400, detail="Conta empresa ainda não foi configurada")

    workplace = await ensure_workplace_access(user, payload.workplaceId, for_management=True)
    memberships = await db.list_enterprise_memberships(enterprise_id, ["accepted"])
    target_membership = next((item for item in memberships if item.get("user_id") == payload.employeeUserId), None)
    if not target_membership:
        raise HTTPException(status_code=400, detail="Funcionário não pertence a esta empresa")

    existing = await db.find_employee_workplace_assignment(enterprise_id, payload.employeeUserId, workplace["id"])
    if existing:
        return EmployeeWorkplaceAssignmentResponse(
            id=existing["id"],
            enterpriseId=existing["enterprise_id"],
            employeeUserId=existing["employee_user_id"],
            workplaceId=existing["workplace_id"],
            assignedBy=existing["assigned_by"],
            createdAt=parse_db_datetime(existing["created_at"]) or datetime.utcnow(),
        )

    created = await db.create_employee_workplace_assignment(
        {
            "enterprise_id": enterprise_id,
            "employee_user_id": payload.employeeUserId,
            "workplace_id": workplace["id"],
            "assigned_by": user["id"],
            "created_at": datetime.utcnow().isoformat(),
        }
    )
    if not created:
        raise HTTPException(status_code=500, detail="Erro ao atribuir local de trabalho")

    return EmployeeWorkplaceAssignmentResponse(
        id=created["id"],
        enterpriseId=created["enterprise_id"],
        employeeUserId=created["employee_user_id"],
        workplaceId=created["workplace_id"],
        assignedBy=created["assigned_by"],
        createdAt=parse_db_datetime(created["created_at"]) or datetime.utcnow(),
    )


@api_router.delete("/enterprise/workplace-assignments/{assignment_id}")
async def remove_enterprise_workplace_assignment(assignment_id: str, user=Depends(get_current_user)):
    require_enterprise_owner(user)
    assignments = await db.list_employee_workplace_assignments(user["enterprise_id"])
    assignment = next((item for item in assignments if item["id"] == assignment_id), None)
    if not assignment:
        raise HTTPException(status_code=404, detail="Atribuição não encontrada")

    deleted = await db.delete_employee_workplace_assignment(assignment_id)
    if not deleted:
        raise HTTPException(status_code=500, detail="Erro ao remover atribuição")

    target_profile = await db.find_profile_by_id(assignment["employee_user_id"])
    if target_profile and target_profile.get("active_workplace_id") == assignment["workplace_id"]:
        await db.update_profile(assignment["employee_user_id"], {"active_workplace_id": None})

    return {"message": "Atribuição removida com sucesso"}


@api_router.get("/enterprise/timesheets/{employee_user_id}", response_model=List[DayTimesheetResponse])
async def get_enterprise_employee_timesheet(
    employee_user_id: str,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    require_enterprise_owner(user)
    enterprise_id = user.get("enterprise_id")
    if not enterprise_id:
        return []

    memberships = await db.list_enterprise_memberships(enterprise_id, ["accepted"])
    membership = next((item for item in memberships if item.get("user_id") == employee_user_id), None)
    if not membership:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado na empresa")

    accepted_at = parse_db_datetime(membership.get("accepted_at"))
    accepted_date_str = accepted_at.strftime("%Y-%m-%d") if accepted_at else None

    if not from_date:
        from_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not to_date:
        to_date = datetime.utcnow().strftime("%Y-%m-%d")
    if accepted_date_str and from_date < accepted_date_str:
        from_date = accepted_date_str

    workplaces = await db.find_employee_assigned_workplaces(enterprise_id, employee_user_id)
    return await build_user_timesheet(employee_user_id, workplaces, from_date, to_date)


@api_router.get("/workplaces", response_model=List[WorkplaceResponse])
async def list_workplaces(user=Depends(get_current_user)):
    workplaces = await get_accessible_workplaces(user)
    active_workplace_id = user.get("active_workplace_id")

    assigned_count_by_workplace: Dict[str, int] = {}
    if user.get("role") == "enterprise_owner" and user.get("enterprise_id"):
        assignments = await db.list_employee_workplace_assignments(user["enterprise_id"])
        for assignment in assignments:
            assigned_count_by_workplace[assignment["workplace_id"]] = assigned_count_by_workplace.get(assignment["workplace_id"], 0) + 1

    return [
        normalize_workplace_response(
            workplace,
            active_workplace_id=active_workplace_id,
            assignment=workplace.get("assignment"),
            assigned_employee_count=assigned_count_by_workplace.get(workplace["id"]),
        )
        for workplace in workplaces
    ]


@api_router.post("/workplaces", response_model=WorkplaceResponse)
async def create_workplace(workplace: WorkplaceCreate, user=Depends(get_current_user)):
    if user.get("role") == "employee" and user.get("enterprise_id"):
        raise HTTPException(status_code=403, detail="Funcionário associado não pode criar locais de trabalho")

    if user.get("role") == "enterprise_owner" and not user.get("enterprise_id"):
        raise HTTPException(status_code=400, detail="Conta empresa ainda não foi configurada")

    now_iso = datetime.utcnow().isoformat()
    workplace_doc = {
        "user_id": user["id"] if user.get("role") != "enterprise_owner" else None,
        "enterprise_id": user.get("enterprise_id") if user.get("role") == "enterprise_owner" else None,
        "name": workplace.name,
        "latitude": workplace.latitude,
        "longitude": workplace.longitude,
        "radius_meters": workplace.radiusMeters,
        "workdays": workplace.workdays.dict(),
        "schedule": workplace.schedule.dict() if workplace.schedule else {"startTime": "09:00", "endTime": "18:00", "marginMinutes": 120},
        "location_locked": True,
        "configured_at": now_iso,
        "created_at": now_iso,
    }

    created = await db.create_workplace(workplace_doc)
    if not created:
        raise HTTPException(status_code=500, detail="Erro ao criar local de trabalho")

    is_active = False
    if user.get("role") != "enterprise_owner" and not user.get("enterprise_id"):
        count = await db.count_personal_workplaces(user["id"])
        is_active = count == 1
        if is_active:
            await db.update_profile(user["id"], {"active_workplace_id": created["id"]})

    return normalize_workplace_response(
        created,
        active_workplace_id=created["id"] if is_active else user.get("active_workplace_id"),
    )


@api_router.put("/workplaces/{workplace_id}", response_model=WorkplaceResponse)
async def update_workplace(workplace_id: str, update: WorkplaceUpdate, user=Depends(get_current_user)):
    workplace = await ensure_workplace_access(user, workplace_id, for_management=True)
    if user.get("role") == "employee" and user.get("enterprise_id"):
        raise HTTPException(status_code=403, detail="Funcionário não pode editar locais de trabalho")

    update_dict: Dict[str, Any] = {}
    if update.name is not None:
        update_dict["name"] = update.name.strip()
    if update.radiusMeters is not None:
        update_dict["radius_meters"] = update.radiusMeters
    if update.workdays is not None:
        update_dict["workdays"] = update.workdays.dict()
    if update.schedule is not None:
        update_dict["schedule"] = update.schedule.dict()

    if update_dict:
        updated = await db.update_workplace(workplace_id, update_dict)
        if not updated:
            raise HTTPException(status_code=500, detail="Erro ao atualizar local de trabalho")
        workplace = updated

    assignment = workplace.get("assignment")
    assigned_employee_count = None
    if user.get("role") == "enterprise_owner" and workplace.get("enterprise_id"):
        assignments = await db.list_employee_workplace_assignments(workplace["enterprise_id"])
        assigned_employee_count = sum(1 for item in assignments if item["workplace_id"] == workplace["id"])

    return normalize_workplace_response(
        workplace,
        active_workplace_id=user.get("active_workplace_id"),
        assignment=assignment,
        assigned_employee_count=assigned_employee_count,
    )


@api_router.post("/workplaces/{workplace_id}/activate")
async def set_active_workplace(workplace_id: str, user=Depends(get_current_user)):
    if user.get("role") == "enterprise_owner":
        raise HTTPException(status_code=403, detail="Conta empresa não usa local ativo")

    workplace = await ensure_workplace_access(user, workplace_id)
    await db.update_profile(user["id"], {"active_workplace_id": workplace_id})
    return {"message": f"'{workplace['name']}' definido como local de trabalho ativo"}


@api_router.get("/workplaces/active", response_model=Optional[WorkplaceResponse])
async def get_active_workplace(user=Depends(get_current_user)):
    if user.get("role") == "enterprise_owner":
        return None

    active_workplace_id = user.get("active_workplace_id")
    if not active_workplace_id:
        return None

    try:
        workplace = await ensure_workplace_access(user, active_workplace_id)
    except HTTPException:
        return None

    return normalize_workplace_response(
        workplace,
        active_workplace_id=active_workplace_id,
        assignment=workplace.get("assignment"),
    )


@api_router.get("/workplace", response_model=Optional[WorkplaceResponse])
async def get_user_workplace(user=Depends(get_current_user)):
    return await get_active_workplace(user)


@api_router.post("/punch", response_model=PunchResponse)
async def create_punch(punch: PunchCreate, user=Depends(get_current_user)):
    require_worker_account(user)

    active_workplace_id = user.get("active_workplace_id")
    if not active_workplace_id:
        raise HTTPException(status_code=400, detail="Nenhum local de trabalho ativo. Configure um local primeiro.")

    workplace = await ensure_workplace_access(user, active_workplace_id)
    server_time = datetime.utcnow()
    device_time = punch.deviceTime or server_time
    today = server_time.strftime("%Y-%m-%d")

    distance = calculate_distance(
        punch.latitude,
        punch.longitude,
        float(workplace["latitude"]),
        float(workplace["longitude"]),
    )
    outside_workplace = distance > workplace["radius_meters"]

    today_punches = await db.find_punches({"user_id": user["id"], "date": today})
    punch_types_today = [item["punch_type"] for item in today_punches]

    if punch.punchType == "IN" and "IN" in punch_types_today:
        raise HTTPException(status_code=400, detail="Já existe um registo de entrada hoje")
    if punch.punchType == "OUT" and "IN" not in punch_types_today:
        raise HTTPException(status_code=400, detail="Não pode registar saída sem entrada")
    if punch.punchType == "OUT" and "OUT" in punch_types_today:
        raise HTTPException(status_code=400, detail="Já existe um registo de saída hoje")
    if punch.punchType == "BREAK_START" and "IN" not in punch_types_today:
        raise HTTPException(status_code=400, detail="Não pode iniciar pausa sem entrada")
    if punch.punchType == "BREAK_START" and punch_types_today.count("BREAK_START") > punch_types_today.count("BREAK_END"):
        raise HTTPException(status_code=400, detail="Já existe uma pausa em curso")
    if punch.punchType == "BREAK_END" and punch_types_today.count("BREAK_START") <= punch_types_today.count("BREAK_END"):
        raise HTTPException(status_code=400, detail="Não existe uma pausa em curso")

    punch_doc = {
        "user_id": user["id"],
        "workplace_id": workplace["id"],
        "workplace_name": workplace["name"],
        "date": today,
        "punch_type": punch.punchType,
        "occurred_at": device_time.isoformat(),
        "received_at": server_time.isoformat(),
        "latitude": punch.latitude,
        "longitude": punch.longitude,
        "accuracy_meters": punch.accuracy,
        "distance_to_workplace_meters": distance,
        "method": punch.method,
        "outside_workplace": outside_workplace,
        "note": punch.note,
    }

    punch_id = await db.create_punch(punch_doc)
    if not punch_id:
        raise HTTPException(status_code=500, detail="Erro ao criar registo de ponto")

    return PunchResponse(
        id=punch_id,
        userId=user["id"],
        workplaceId=workplace["id"],
        workplaceName=workplace["name"],
        date=today,
        punchType=punch.punchType,
        occurredAt=device_time,
        receivedAt=server_time,
        latitude=punch.latitude,
        longitude=punch.longitude,
        accuracyMeters=punch.accuracy,
        distanceToWorkplaceMeters=distance,
        method=punch.method,
        outsideWorkplace=outside_workplace,
        note=punch.note,
    )


@api_router.post("/punch/manual")
async def manual_punch_legacy(punch: PunchCreate, user=Depends(get_current_user)):
    return await create_punch(punch, user)


@api_router.post("/break/manual")
async def manual_break_legacy(break_data: PunchCreate, user=Depends(get_current_user)):
    return await create_punch(break_data, user)


@api_router.post("/events/geofence")
async def create_geofence_event(event: GeofenceEventCreate, user=Depends(get_current_user)):
    require_worker_account(user)

    active_workplace_id = user.get("active_workplace_id")
    if not active_workplace_id:
        raise HTTPException(status_code=400, detail="Nenhum local de trabalho ativo")

    workplace = await ensure_workplace_access(user, active_workplace_id)
    existing = await db.find_geofence_event(event.eventId, user["id"])
    if existing:
        return {"message": "Evento já processado"}

    created = await db.create_geofence_event(
        {
            "event_id": event.eventId,
            "user_id": user["id"],
            "workplace_id": workplace["id"],
            "event_type": event.eventType,
            "latitude": event.latitude,
            "longitude": event.longitude,
            "accuracy": event.accuracy,
            "device_time": event.deviceTime.isoformat() if event.deviceTime else None,
            "received_at": datetime.utcnow().isoformat(),
            "processed": False,
        }
    )
    if not created:
        raise HTTPException(status_code=500, detail="Erro ao guardar evento de geofence")

    distance = calculate_distance(
        event.latitude,
        event.longitude,
        float(workplace["latitude"]),
        float(workplace["longitude"]),
    )
    message = "Dentro do local de trabalho" if distance <= workplace["radius_meters"] else "Fora do local de trabalho"
    return {
        "eventId": event.eventId,
        "eventType": event.eventType,
        "workplaceName": workplace["name"],
        "distance": int(distance),
        "message": message,
    }


@api_router.get("/timesheet", response_model=List[DayTimesheetResponse])
async def get_timesheet(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    require_worker_account(user)

    if not from_date:
        from_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not to_date:
        to_date = datetime.utcnow().strftime("%Y-%m-%d")

    workplaces = await get_accessible_workplaces(user)
    return await build_user_timesheet(user["id"], workplaces, from_date, to_date)


@api_router.get("/timesheet/today")
async def get_today_status(user=Depends(get_current_user)):
    require_worker_account(user)
    return await build_today_status(user)


@api_router.get("/export/timesheet.csv")
async def export_csv(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    require_worker_account(user)

    if not from_date:
        from_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not to_date:
        to_date = datetime.utcnow().strftime("%Y-%m-%d")

    workplaces = await get_accessible_workplaces(user)
    workplace_map = {item["id"]: item for item in workplaces}
    workplace_ids = list(workplace_map.keys())
    punches_data = await db.find_punches_by_user_date_range(user["id"], from_date, to_date, workplace_ids or None)
    days = group_punches_by_day(punches_data)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Data", "Local de Trabalho", "Dias de Trabalho Config.", "Link Mapa Local",
        "Entrada", "Método Entrada", "Fora Local Entrada", "Distância Entrada (m)", "Precisão Entrada (m)", "Link Mapa Entrada",
        "Saída", "Método Saída", "Fora Local Saída", "Distância Saída (m)", "Precisão Saída (m)", "Link Mapa Saída",
        "Pausas (min)", "Bruto (min)", "Líquido (min)", "Horas Trabalhadas", "Notas"
    ])

    for date_str in sorted(days.keys()):
        day_data = days[date_str]
        workplace = workplace_map.get(day_data["workplaceId"], {})
        workdays_str = ""
        if workplace.get("workdays"):
            labels = [("monday", "Seg"), ("tuesday", "Ter"), ("wednesday", "Qua"), ("thursday", "Qui"), ("friday", "Sex"), ("saturday", "Sáb"), ("sunday", "Dom")]
            workdays_str = ", ".join(label for key, label in labels if workplace["workdays"].get(key))

        punch_in = next((item for item in day_data["punches"] if item["type"] == "IN"), None)
        punch_out = next((item for item in reversed(day_data["punches"]) if item["type"] == "OUT"), None)
        break_minutes = 0
        notes: List[str] = []
        break_start = None

        for punch in day_data["punches"]:
            if punch["type"] == "BREAK_START":
                break_start = punch
            elif punch["type"] == "BREAK_END" and break_start:
                break_minutes += int((punch["occurredAt"] - break_start["occurredAt"]).total_seconds() / 60)
                break_start = None
            if punch.get("note"):
                notes.append(f"{punch['type']}: {punch['note']}")

        gross_minutes = 0
        if punch_in and punch_out:
            gross_minutes = int((punch_out["occurredAt"] - punch_in["occurredAt"]).total_seconds() / 60)
        net_minutes = max(0, gross_minutes - break_minutes)

        writer.writerow([
            date_str,
            day_data["workplaceName"],
            workdays_str,
            generate_maps_link(float(workplace["latitude"]), float(workplace["longitude"])) if workplace else "",
            punch_in["occurredAt"].strftime("%H:%M:%S") if punch_in else "",
            punch_in["method"] if punch_in else "",
            "Sim" if punch_in and punch_in["outsideWorkplace"] else "Não" if punch_in else "",
            int(punch_in["distance"]) if punch_in else "",
            int(punch_in["accuracy"]) if punch_in else "",
            punch_in["mapsLink"] if punch_in else "",
            punch_out["occurredAt"].strftime("%H:%M:%S") if punch_out else "",
            punch_out["method"] if punch_out else "",
            "Sim" if punch_out and punch_out["outsideWorkplace"] else "Não" if punch_out else "",
            int(punch_out["distance"]) if punch_out else "",
            int(punch_out["accuracy"]) if punch_out else "",
            punch_out["mapsLink"] if punch_out else "",
            break_minutes,
            gross_minutes,
            net_minutes,
            format_minutes(net_minutes),
            "; ".join(notes),
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=folha_ponto_{from_date}_{to_date}.csv"},
    )


@api_router.get("/export/timesheet.xlsx")
async def export_xlsx(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    user=Depends(get_current_user),
):
    require_worker_account(user)

    if not from_date:
        from_date = (datetime.utcnow() - timedelta(days=30)).strftime("%Y-%m-%d")
    if not to_date:
        to_date = datetime.utcnow().strftime("%Y-%m-%d")

    workplaces = await get_accessible_workplaces(user)
    workplace_map = {item["id"]: item for item in workplaces}
    workplace_ids = list(workplace_map.keys())
    punches_data = await db.find_punches_by_user_date_range(user["id"], from_date, to_date, workplace_ids or None)
    days = group_punches_by_day(punches_data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Folha de Ponto"
    headers = ["Data", "Local de Trabalho", "Dias Config.", "Entrada", "Saída", "Pausas", "Bruto", "Líquido", "Fora Local", "Notas"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    total_net = 0
    for date_str in sorted(days.keys()):
        day_data = days[date_str]
        workplace = workplace_map.get(day_data["workplaceId"], {})
        workdays_str = ""
        if workplace.get("workdays"):
            labels = [("monday", "Seg"), ("tuesday", "Ter"), ("wednesday", "Qua"), ("thursday", "Qui"), ("friday", "Sex"), ("saturday", "Sáb"), ("sunday", "Dom")]
            workdays_str = ", ".join(label for key, label in labels if workplace["workdays"].get(key))

        punch_in = next((item for item in day_data["punches"] if item["type"] == "IN"), None)
        punch_out = next((item for item in reversed(day_data["punches"]) if item["type"] == "OUT"), None)
        break_minutes = 0
        outside_flags: List[str] = []
        notes: List[str] = []
        break_start = None

        for punch in day_data["punches"]:
            if punch["type"] == "IN" and punch["outsideWorkplace"]:
                outside_flags.append("Entrada")
            if punch["type"] == "OUT" and punch["outsideWorkplace"]:
                outside_flags.append("Saída")
            if punch["type"] == "BREAK_START":
                break_start = punch
            elif punch["type"] == "BREAK_END" and break_start:
                break_minutes += int((punch["occurredAt"] - break_start["occurredAt"]).total_seconds() / 60)
                break_start = None
            if punch.get("note"):
                notes.append(punch["note"])

        gross_minutes = 0
        if punch_in and punch_out:
            gross_minutes = int((punch_out["occurredAt"] - punch_in["occurredAt"]).total_seconds() / 60)
        net_minutes = max(0, gross_minutes - break_minutes)
        total_net += net_minutes

        ws.append([
            date_str,
            day_data["workplaceName"],
            workdays_str,
            punch_in["occurredAt"].strftime("%H:%M") if punch_in else "-",
            punch_out["occurredAt"].strftime("%H:%M") if punch_out else "-",
            format_minutes(break_minutes),
            format_minutes(gross_minutes),
            format_minutes(net_minutes),
            ", ".join(outside_flags) if outside_flags else "-",
            "; ".join(notes) if notes else "",
        ])

    ws.append([])
    ws.append(["", "", "", "", "", "", "TOTAL:", format_minutes(total_net), "", ""])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max(14, max_length + 2)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=folha_ponto_{from_date}_{to_date}.xlsx"},
    )


@api_router.get("/")
async def root():
    return {"message": "GeoPunch API", "version": "4.0.0"}


@api_router.get("/health")
async def health():
    return {"status": "ok"}


@api_router.post("/auth/register")
async def legacy_register():
    return JSONResponse(
        status_code=410,
        content={"message": "Este endpoint é legado.", "info": "Use Supabase Auth diretamente no cliente."},
    )


@api_router.post("/auth/login")
async def legacy_login():
    return JSONResponse(
        status_code=410,
        content={"message": "Este endpoint é legado.", "info": "Use Supabase Auth diretamente no cliente."},
    )


@api_router.post("/seed")
async def seed_data():
    return {
        "message": "Seed endpoint deprecated.",
        "info": "Please use Supabase Auth to register users. Profiles are created automatically.",
    }


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
